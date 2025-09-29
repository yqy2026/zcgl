"""
租金合同Excel导入导出服务
支持合同和台账数据的批量导入导出
"""

import polars as pl
from typing import Dict, Any, List, Optional, Union, Tuple
from pathlib import Path
import logging
from datetime import datetime, date, timedelta
import tempfile
import os
import uuid
from decimal import Decimal
from sqlalchemy.orm import Session
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.models.rent_contract import RentContract, RentTerm, RentLedger
from src.models.asset import Asset
from src.models import Ownership
from src.crud.rent_contract import rent_contract, rent_term, rent_ledger
from src.crud.asset import asset_crud
from src.crud.ownership import ownership
from src.schemas.rent_contract import RentContractCreate, RentTermCreate, RentLedgerCreate
from src.database import get_db

logger = logging.getLogger(__name__)


class RentContractExcelError(Exception):
    """租金合同Excel操作异常"""
    pass


class RentContractExcelService:
    """租金合同Excel导入导出服务"""

    # 合同Excel列名映射
    CONTRACT_COLUMN_MAPPING = {
        "合同编号": "contract_number",
        "资产名称": "asset_name",
        "资产地址": "asset_address",
        "权属方名称": "ownership_name",
        "承租方名称": "tenant_name",
        "承租方联系方式": "tenant_contact",
        "承租方证件类型": "tenant_id_type",
        "承租方证件号码": "tenant_id_number",
        "合同开始日期": "start_date",
        "合同结束日期": "end_date",
        "合同总金额": "total_amount",
        "支付方式": "payment_method",
        "押金金额": "deposit_amount",
        "押金支付方式": "deposit_payment_method",
        "违约金比例": "penalty_rate",
        "合同状态": "contract_status",
        "备注": "notes"
    }

    # 租金条款Excel列名映射
    TERM_COLUMN_MAPPING = {
        "合同编号": "contract_number",
        "开始日期": "start_date",
        "结束日期": "end_date",
        "月租金": "monthly_rent",
        "支付方式": "payment_method",
        "支付周期": "payment_cycle",
        "备注": "notes"
    }

    # 台账Excel列名映射
    LEDGER_COLUMN_MAPPING = {
        "合同编号": "contract_number",
        "年月": "year_month",
        "应收金额": "due_amount",
        "实收金额": "paid_amount",
        "欠款金额": "overdue_amount",
        "支付状态": "payment_status",
        "支付日期": "payment_date",
        "备注": "notes"
    }

    # 必填字段
    CONTRACT_REQUIRED_FIELDS = [
        "合同编号", "资产名称", "权属方名称", "承租方名称",
        "合同开始日期", "合同结束日期", "月租金"
    ]

    TERM_REQUIRED_FIELDS = [
        "合同编号", "开始日期", "结束日期", "月租金"
    ]

    def __init__(self):
        self.db = next(get_db())

    def export_contracts_to_excel(
        self,
        contract_ids: Optional[List[str]] = None,
        include_terms: bool = True,
        include_ledger: bool = True,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> Dict[str, Any]:
        """导出租金合同数据到Excel文件"""
        try:
            # 获取合同数据
            contracts = self._get_contracts_for_export(
                contract_ids, start_date, end_date
            )

            if not contracts:
                return {
                    "success": False,
                    "message": "没有找到符合条件的合同数据",
                    "file_path": None
                }

            # 创建Excel文件
            temp_dir = tempfile.gettempdir()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"租金合同导出_{timestamp}.xlsx"
            file_path = os.path.join(temp_dir, filename)

            # 创建工作簿
            from openpyxl import Workbook
            wb = Workbook()

            # 删除默认工作表
            wb.remove(wb.active)

            # 导出合同数据
            self._export_contracts_sheet(wb, contracts)

            # 导出租金条款数据
            if include_terms:
                self._export_terms_sheet(wb, contracts)

            # 导出台账数据
            if include_ledger:
                self._export_ledger_sheet(wb, contracts)

            # 保存文件
            wb.save(file_path)

            # 获取文件大小
            file_size = os.path.getsize(file_path)

            logger.info(f"成功导出 {len(contracts)} 个合同到 {file_path}")

            return {
                "success": True,
                "message": f"成功导出 {len(contracts)} 个合同数据",
                "file_path": file_path,
                "file_name": filename,
                "file_size": file_size,
                "stats": {
                    "total_contracts": len(contracts),
                    "total_terms": sum(len(c.terms) for c in contracts),
                    "total_ledgers": sum(len(c.ledgers) for c in contracts)
                }
            }

        except Exception as e:
            logger.error(f"导出合同Excel失败: {str(e)}")
            return {
                "success": False,
                "message": f"导出失败: {str(e)}",
                "file_path": None
            }

    def import_contracts_from_excel(
        self,
        file_path: str,
        import_terms: bool = True,
        import_ledger: bool = False,
        overwrite_existing: bool = False
    ) -> Dict[str, Any]:
        """从Excel文件导入租金合同数据"""
        try:
            if not os.path.exists(file_path):
                raise RentContractExcelError("文件不存在")

            # 加载Excel文件
            wb = load_workbook(file_path, data_only=True)

            results = {
                "success": True,
                "message": "导入成功",
                "imported_contracts": 0,
                "imported_terms": 0,
                "imported_ledgers": 0,
                "errors": [],
                "warnings": []
            }

            # 导入合同数据
            if "合同信息" in wb.sheetnames:
                contract_result = self._import_contracts_sheet(
                    wb["合同信息"],
                    overwrite_existing
                )
                results.update(contract_result)

            # 导入租金条款数据
            if import_terms and "租金条款" in wb.sheetnames:
                terms_result = self._import_terms_sheet(wb["租金条款"])
                results.update(terms_result)

            # 导入台账数据
            if import_ledger and "租金台账" in wb.sheetnames:
                ledger_result = self._import_ledger_sheet(wb["租金台账"])
                results.update(ledger_result)

            logger.info(f"Excel导入完成: {results}")
            return results

        except Exception as e:
            logger.error(f"导入合同Excel失败: {str(e)}")
            return {
                "success": False,
                "message": f"导入失败: {str(e)}",
                "imported_contracts": 0,
                "errors": [str(e)]
            }

    def download_contract_template(self) -> Dict[str, Any]:
        """下载合同导入模板"""
        try:
            # 创建模板文件
            temp_dir = tempfile.gettempdir()
            filename = "租金合同导入模板.xlsx"
            file_path = os.path.join(temp_dir, filename)

            from openpyxl import Workbook
            wb = Workbook()

            # 删除默认工作表
            wb.remove(wb.active)

            # 创建模板工作表
            self._create_contract_template(wb)
            self._create_terms_template(wb)
            self._create_ledger_template(wb)

            # 保存文件
            wb.save(file_path)

            file_size = os.path.getsize(file_path)

            return {
                "success": True,
                "message": "模板下载成功",
                "file_path": file_path,
                "file_name": filename,
                "file_size": file_size
            }

        except Exception as e:
            logger.error(f"下载模板失败: {str(e)}")
            return {
                "success": False,
                "message": f"下载模板失败: {str(e)}",
                "file_path": None
            }

    def _get_contracts_for_export(
        self,
        contract_ids: Optional[List[str]] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> List[RentContract]:
        """获取要导出的合同数据 - 优化版本使用join预加载关联数据"""
        try:
            # 使用join预加载关联数据，避免N+1查询
            query = self.db.query(RentContract).join(
                Asset, RentContract.asset_id == Asset.id, isouter=True
            ).join(
                Ownership, RentContract.ownership_id == Ownership.id, isouter=True
            )

            if contract_ids:
                query = query.filter(RentContract.id.in_(contract_ids))

            if start_date:
                query = query.filter(RentContract.start_date >= start_date)

            if end_date:
                query = query.filter(RentContract.end_date <= end_date)

            contracts = query.all()

            return contracts

        except Exception as e:
            logger.error(f"获取合同数据失败: {str(e)}")
            raise RentContractExcelError(f"获取合同数据失败: {str(e)}")

    def _export_contracts_sheet(self, wb: Workbook, contracts: List[RentContract]):
        """导出合同信息工作表 - 优化版本使用预加载的关联数据"""
        ws = wb.create_sheet("合同信息")

        # 设置表头
        headers = list(self.CONTRACT_COLUMN_MAPPING.keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # 填充数据
        for row, contract in enumerate(contracts, 2):
            for col, header in enumerate(headers, 1):
                field_name = self.CONTRACT_COLUMN_MAPPING[header]

                # 处理关联数据字段
                if header == "资产名称":
                    value = contract.asset.property_name if contract.asset else ""
                elif header == "资产地址":
                    value = contract.asset.address if contract.asset else ""
                elif header == "权属方名称":
                    value = contract.ownership.ownership_name if contract.ownership else ""
                else:
                    value = getattr(contract, field_name, "")

                # 处理日期类型
                if isinstance(value, (date, datetime)):
                    value = value.strftime("%Y-%m-%d")
                elif isinstance(value, Decimal):
                    value = float(value)

                ws.cell(row=row, column=col, value=value)

        # 设置列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _export_terms_sheet(self, wb: Workbook, contracts: List[RentContract]):
        """导出租金条款工作表"""
        ws = wb.create_sheet("租金条款")

        # 设置表头
        headers = list(self.TERM_COLUMN_MAPPING.keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # 填充数据
        row = 2
        for contract in contracts:
            if hasattr(contract, 'terms') and contract.terms:
                for term in contract.terms:
                    for col, header in enumerate(headers, 1):
                        field_name = self.TERM_COLUMN_MAPPING[header]
                        value = getattr(term, field_name, "")

                        # 处理特殊字段
                        if header == "合同编号":
                            value = contract.contract_number
                        elif isinstance(value, (date, datetime)):
                            value = value.strftime("%Y-%m-%d")
                        elif isinstance(value, Decimal):
                            value = float(value)

                        ws.cell(row=row, column=col, value=value)
                    row += 1

        # 设置列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _export_ledger_sheet(self, wb: Workbook, contracts: List[RentContract]):
        """导出台账工作表"""
        ws = wb.create_sheet("租金台账")

        # 设置表头
        headers = list(self.LEDGER_COLUMN_MAPPING.keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # 填充数据
        row = 2
        for contract in contracts:
            if hasattr(contract, 'ledgers') and contract.ledgers:
                for ledger in contract.ledgers:
                    for col, header in enumerate(headers, 1):
                        field_name = self.LEDGER_COLUMN_MAPPING[header]
                        value = getattr(ledger, field_name, "")

                        # 处理特殊字段
                        if header == "合同编号":
                            value = contract.contract_number
                        elif isinstance(value, (date, datetime)):
                            value = value.strftime("%Y-%m-%d")
                        elif isinstance(value, Decimal):
                            value = float(value)

                        ws.cell(row=row, column=col, value=value)
                    row += 1

        # 设置列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_contract_template(self, wb: Workbook):
        """创建合同信息模板"""
        ws = wb.create_sheet("合同信息")

        # 设置表头
        headers = list(self.CONTRACT_COLUMN_MAPPING.keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # 添加示例数据
        example_data = {
            "合同编号": "HT2024001",
            "资产名称": "示例资产",
            "资产地址": "示例地址",
            "权属方名称": "示例权属方",
            "承租方名称": "示例承租方",
            "承租方联系方式": "13800138000",
            "承租方证件类型": "营业执照",
            "承租方证件号码": "123456789",
            "合同开始日期": "2024-01-01",
            "合同结束日期": "2024-12-31",
            "合同总金额": "120000",
            "支付方式": "银行转账",
            "押金金额": "10000",
            "押金支付方式": "银行转账",
            "违约金比例": "0.001",
            "合同状态": "生效",
            "备注": "示例备注"
        }

        for col, header in enumerate(headers, 1):
            ws.cell(row=2, column=col, value=example_data.get(header, ""))

        # 设置列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_terms_template(self, wb: Workbook):
        """创建租金条款模板"""
        ws = wb.create_sheet("租金条款")

        # 设置表头
        headers = list(self.TERM_COLUMN_MAPPING.keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # 添加示例数据
        example_data = {
            "合同编号": "HT2024001",
            "开始日期": "2024-01-01",
            "结束日期": "2024-06-30",
            "月租金": "10000",
            "支付方式": "银行转账",
            "支付周期": "月付",
            "备注": "上半年租金"
        }

        for col, header in enumerate(headers, 1):
            ws.cell(row=2, column=col, value=example_data.get(header, ""))

        # 设置列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_ledger_template(self, wb: Workbook):
        """创建台账模板"""
        ws = wb.create_sheet("租金台账")

        # 设置表头
        headers = list(self.LEDGER_COLUMN_MAPPING.keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # 添加示例数据
        example_data = {
            "合同编号": "HT2024001",
            "年月": "2024-01",
            "应收金额": "10000",
            "实收金额": "10000",
            "欠款金额": "0",
            "支付状态": "已支付",
            "支付日期": "2024-01-05",
            "备注": "1月份租金"
        }

        for col, header in enumerate(headers, 1):
            ws.cell(row=2, column=col, value=example_data.get(header, ""))

        # 设置列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _import_contracts_sheet(
        self,
        ws: Worksheet,
        overwrite_existing: bool
    ) -> Dict[str, Any]:
        """导入合同信息工作表"""
        result = {
            "imported_contracts": 0,
            "errors": [],
            "warnings": []
        }

        try:
            # 读取表头
            headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    headers.append(header)

            # 验证必填字段
            missing_fields = set(self.CONTRACT_REQUIRED_FIELDS) - set(headers)
            if missing_fields:
                raise RentContractExcelError(f"缺少必填字段: {', '.join(missing_fields)}")

            # 处理数据行
            for row in range(2, ws.max_row + 1):
                try:
                    row_data = {}
                    for col, header in enumerate(headers, 1):
                        cell_value = ws.cell(row=row, column=col).value
                        row_data[header] = cell_value

                    # 验证必填字段
                    if not all(row_data.get(field) for field in self.CONTRACT_REQUIRED_FIELDS):
                        result["warnings"].append(f"第{row}行: 缺少必填字段，跳过")
                        continue

                    # 查找资产和权属方
                    asset = asset_crud.get_by_property_name(
                        self.db,
                        row_data["资产名称"]
                    )
                    if not asset:
                        result["warnings"].append(f"第{row}行: 未找到资产'{row_data['资产名称']}'，跳过")
                        continue

                    ownership_entity = ownership.get_by_name(
                        self.db,
                        row_data["权属方名称"]
                    )
                    if not ownership_entity:
                        result["warnings"].append(f"第{row}行: 未找到权属方'{row_data['权属方名称']}'，跳过")
                        continue

                    # 检查合同编号是否重复
                    existing_contract = rent_contract.get_by_contract_number(
                        self.db,
                        row_data["合同编号"]
                    )
                    if existing_contract and not overwrite_existing:
                        result["warnings"].append(f"第{row}行: 合同编号'{row_data['合同编号']}'已存在，跳过")
                        continue

                    # 创建合同数据
                    contract_data = {
                        "contract_number": row_data["合同编号"],
                        "asset_id": asset.id,
                        "ownership_id": ownership_entity.id,
                        "tenant_name": row_data["承租方名称"],
                        "tenant_contact": row_data.get("承租方联系方式"),
                        "tenant_id_type": row_data.get("承租方证件类型"),
                        "tenant_id_number": row_data.get("承租方证件号码"),
                        "start_date": self._parse_date(row_data["合同开始日期"]),
                        "end_date": self._parse_date(row_data["合同结束日期"]),
                        "total_amount": Decimal(str(row_data.get("合同总金额", 0))),
                        "payment_method": row_data.get("支付方式"),
                        "deposit_amount": Decimal(str(row_data.get("押金金额", 0))),
                        "deposit_payment_method": row_data.get("押金支付方式"),
                        "penalty_rate": Decimal(str(row_data.get("违约金比例", 0))),
                        "contract_status": row_data.get("合同状态", "草稿"),
                        "notes": row_data.get("备注")
                    }

                    # 创建或更新合同
                    if existing_contract and overwrite_existing:
                        contract = rent_contract.update(
                            self.db,
                            db_obj=existing_contract,
                            obj_in=contract_data
                        )
                    else:
                        contract = rent_contract.create(self.db, obj_in=contract_data)

                    result["imported_contracts"] += 1

                except Exception as e:
                    result["errors"].append(f"第{row}行: {str(e)}")
                    continue

            return result

        except Exception as e:
            logger.error(f"导入合同信息失败: {str(e)}")
            result["errors"].append(f"导入合同信息失败: {str(e)}")
            return result

    def _import_terms_sheet(self, ws: Worksheet) -> Dict[str, Any]:
        """导入租金条款工作表"""
        result = {
            "imported_terms": 0,
            "errors": [],
            "warnings": []
        }

        try:
            # 读取表头
            headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    headers.append(header)

            # 验证必填字段
            missing_fields = set(self.TERM_REQUIRED_FIELDS) - set(headers)
            if missing_fields:
                raise RentContractExcelError(f"缺少必填字段: {', '.join(missing_fields)}")

            # 处理数据行
            for row in range(2, ws.max_row + 1):
                try:
                    row_data = {}
                    for col, header in enumerate(headers, 1):
                        cell_value = ws.cell(row=row, column=col).value
                        row_data[header] = cell_value

                    # 验证必填字段
                    if not all(row_data.get(field) for field in self.TERM_REQUIRED_FIELDS):
                        result["warnings"].append(f"第{row}行: 缺少必填字段，跳过")
                        continue

                    # 查找合同
                    contract = rent_contract.get_by_contract_number(
                        self.db,
                        row_data["合同编号"]
                    )
                    if not contract:
                        result["warnings"].append(f"第{row}行: 未找到合同'{row_data['合同编号']}'，跳过")
                        continue

                    # 创建条款数据
                    term_data = {
                        "contract_id": contract.id,
                        "start_date": self._parse_date(row_data["开始日期"]),
                        "end_date": self._parse_date(row_data["结束日期"]),
                        "monthly_rent": Decimal(str(row_data["月租金"])),
                        "payment_method": row_data.get("支付方式"),
                        "payment_cycle": row_data.get("支付周期", "月付"),
                        "notes": row_data.get("备注")
                    }

                    # 创建条款
                    term = rent_term.create(self.db, obj_in=term_data)
                    result["imported_terms"] += 1

                except Exception as e:
                    result["errors"].append(f"第{row}行: {str(e)}")
                    continue

            return result

        except Exception as e:
            logger.error(f"导入租金条款失败: {str(e)}")
            result["errors"].append(f"导入租金条款失败: {str(e)}")
            return result

    def _import_ledger_sheet(self, ws: Worksheet) -> Dict[str, Any]:
        """导入台账工作表"""
        result = {
            "imported_ledgers": 0,
            "errors": [],
            "warnings": []
        }

        try:
            # 读取表头
            headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    headers.append(header)

            # 处理数据行
            for row in range(2, ws.max_row + 1):
                try:
                    row_data = {}
                    for col, header in enumerate(headers, 1):
                        cell_value = ws.cell(row=row, column=col).value
                        row_data[header] = cell_value

                    # 查找合同
                    contract = rent_contract.get_by_contract_number(
                        self.db,
                        row_data["合同编号"]
                    )
                    if not contract:
                        result["warnings"].append(f"第{row}行: 未找到合同'{row_data['合同编号']}'，跳过")
                        continue

                    # 创建台账数据
                    ledger_data = {
                        "contract_id": contract.id,
                        "year_month": row_data["年月"],
                        "due_amount": Decimal(str(row_data.get("应收金额", 0))),
                        "paid_amount": Decimal(str(row_data.get("实收金额", 0))),
                        "overdue_amount": Decimal(str(row_data.get("欠款金额", 0))),
                        "payment_status": row_data.get("支付状态", "未支付"),
                        "payment_date": self._parse_date(row_data.get("支付日期")),
                        "notes": row_data.get("备注")
                    }

                    # 检查是否已存在
                    existing_ledger = rent_ledger.get_by_contract_and_month(
                        self.db,
                        contract.id,
                        row_data["年月"]
                    )

                    if existing_ledger:
                        # 更新现有台账
                        ledger = rent_ledger.update(
                            self.db,
                            db_obj=existing_ledger,
                            obj_in=ledger_data
                        )
                    else:
                        # 创建新台账
                        ledger = rent_ledger.create(self.db, obj_in=ledger_data)

                    result["imported_ledgers"] += 1

                except Exception as e:
                    result["errors"].append(f"第{row}行: {str(e)}")
                    continue

            return result

        except Exception as e:
            logger.error(f"导入台账失败: {str(e)}")
            result["errors"].append(f"导入台账失败: {str(e)}")
            return result

    def _parse_date(self, date_value: Any) -> Optional[date]:
        """解析日期值"""
        if not date_value:
            return None

        if isinstance(date_value, (date, datetime)):
            return date_value.date() if isinstance(date_value, datetime) else date_value

        if isinstance(date_value, str):
            try:
                # 尝试不同格式
                for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日", "%m/%d/%Y"]:
                    try:
                        return datetime.strptime(date_value, fmt).date()
                    except ValueError:
                        continue
            except Exception:
                pass

        raise ValueError(f"无法解析日期: {date_value}")

    def cleanup_file(self, file_path: str) -> bool:
        """清理临时文件"""
        try:
            if os.path.exists(file_path):
                os.unlink(file_path)
                logger.info(f"已清理文件: {file_path}")
                return True
            return False
        except Exception as e:
            logger.warning(f"清理文件失败: {str(e)}")
            return False


# 单例实例
rent_contract_excel_service = RentContractExcelService()